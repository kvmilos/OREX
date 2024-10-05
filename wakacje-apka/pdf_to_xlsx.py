import pdfminer
import pdfminer.high_level
from pdfminer.high_level import extract_pages
import re
import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

fak_NUMER = r'data i miejsce wystawienia\n\n(.*)\n'
fak_DATA_WYSTAWIENIA = r'(\d{4}-\d{2}-\d{2}) .*\ndata i miejsce wystawienia'
fak_KWOTY = r'W tym.*?Brutto(.*?)Razem do'
fak_REZERWACJA = r'Prowizja .* (1\d{6})'


if not os.path.exists('done'):
    os.makedirs('done')

num_pages = 0
text = ''
file_names = []
for file in os.listdir('.'):
    if file.endswith('.pdf'):
        with open(file, 'rb') as f:
            text += pdfminer.high_level.extract_text(f)
        current_num_pages = len(list(extract_pages(file)))
        num_pages += current_num_pages
        file_names.extend([file] * current_num_pages)
        os.rename(file, 'done/' + file)

pages = text.split('\x0c')

current_page_number = 1
previous_file_name = ''
faktury = [['numer', 'data_wystawienia', 'kwota_netto', 'stawka_vat', 'kwota_brutto', 'rez', 'plik', 'strona']]
korekty = [['plik', 'strona']]
for i, page in enumerate(pages):
    if i >= len(file_names):
        continue
    if i > 0:
        try:
            if file_names[i] != file_names[i-1]:
                current_page_number = 1
            else:
                current_page_number += 1
        except IndexError:
            if i < num_pages:
                print('IndexError at', i, ', file ', file_names[i-1], 'page', current_page_number)
    try:
        numer = re.search(fak_NUMER, page).group(1)
        data_wystawienia = re.search(fak_DATA_WYSTAWIENIA, page).group(1)
        rezerwacja = re.search(fak_REZERWACJA, page).group(1)
        kwoty = re.search(fak_KWOTY, page, re.DOTALL).group(1).replace(' ', '').replace('\n', ' ').replace('%', '% ').strip()
        kw2 = [kwota for kwota in kwoty.split(' ') if kwota != '']
        kwota_netto, stawka_vat, kwota_vat, kwota_brutto = kw2
        kwota_netto = float(kwota_netto.replace(',', '.'))
        kwota_brutto = float(kwota_brutto.replace(',', '.'))
        rezerwacja = int(rezerwacja)
        faktura = numer, data_wystawienia, kwota_netto, stawka_vat, kwota_brutto, rezerwacja, file_names[i], current_page_number
        if rezerwacja not in [item[5] for item in faktury]:
            faktury.append(faktura)
    except AttributeError:
        if i != num_pages:
            print(i+1, 'error lub korekta', 'plik:', file_names[i], 'strona: ', current_page_number)
            korekty.append([file_names[i], current_page_number])
    except ValueError:
        faktury.append(['check', 'check', 'check', 'check', 'check', 'check', file_names[i], current_page_number])
        print('you need to check ' + file_names[i], 'page ' + str(current_page_number))


if len(faktury) > 1:
    fak = pd.DataFrame(faktury)
    kor = pd.DataFrame(korekty)
    with pd.ExcelWriter('faktury_wakacje.xlsx', mode='w') as writer:
        fak.to_excel(writer, sheet_name='faktury', index=False, header=False)
        kor.to_excel(writer, sheet_name='korekty', index=False, header=False)
columns = ['pdf_rez', 'pdf_nr_dok', 'pdf_data_dok', 'samo_start_date', 'samo_country', 'pdf_netto', 'pdf_brutto', 'samo_adjustment_nr', 'samo_date*', 'samo_adjustment*', 'pdf_st_vat', 'samo_rez', 'samo_com', 'samo_debt', 'diff']

rows = []

for row in faktury[1:]:
    new_row = {
        'pdf_rez': row[5],
        'pdf_nr_dok': row[0],
        'pdf_data_dok': row[1],
        'pdf_netto': row[2],
        'pdf_brutto': row[4],
        'pdf_st_vat': row[3],
        'samo_start_date': None,
        'samo_country': None,
        'samo_rez': None,
        'samo_com': None,
        'samo_debt': None,
        'samo_adjustment_nr': None,
        'samo_date*': None,
        'samo_adjustment*': None,
        'diff': None
    }
    rows.append(new_row)

new_faktury = pd.DataFrame(rows)
new_faktury = new_faktury[columns]

with pd.ExcelWriter('faktury_wakacje.xlsx', mode='a', engine='openpyxl') as writer:
    new_faktury.to_excel(writer, sheet_name='analiza', index=False)

wb = load_workbook('faktury_wakacje.xlsx')
ws = wb['analiza']

pdf_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
samo_fill = PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid')
diff_fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')

for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(columns)):
    for cell in col:
        if cell.value.startswith('pdf'):
            cell.fill = pdf_fill
        elif cell.value.startswith('samo'):
            cell.fill = samo_fill
        elif cell.value == 'diff':
            cell.fill = diff_fill
        
wb.save('faktury_wakacje.xlsx')

input('Press Enter to exit...')