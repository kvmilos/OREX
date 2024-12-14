import os
import re
import pdfminer
import pdfminer.high_level
from pdfminer.high_level import extract_pages
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import requests
import json

def fetch_patterns():
    url = "https://raw.githubusercontent.com/kvmilos/OREX/refs/heads/main/regex_patterns.json"
    response = requests.get(url)
    if response.status_code == 200:
        return json.loads(response.text)
    else:
        print("Nie udało się pobrać wzorców.")
        return None

patterns = fetch_patterns()
if patterns is None:
    print("Nie udało się pobrać wzorców.")
    exit(1)

fak_NUMER = patterns["wakacje"]["numer_faktury"]
fak_DATA_WYSTAWIENIA = patterns["wakacje"]["data_faktury"]
fak_TERMIN_PLATNOSCI = patterns["wakacje"]["termin_platnosci"]
fak_KWOTY = patterns["wakacje"]["kwoty"]
fak_REZERWACJA = patterns["wakacje"]["rezerwacja"]

os.makedirs('done', exist_ok=True)

faktury = [['numer', 'data_wystawienia', 'termin_platnosci', 'kwota_netto', 'stawka_vat', 'kwota_brutto', 'rez', 'plik', 'strona']]
korekty = [['plik', 'strona']]
num_pages = 0
text = ''
file_names = []

for file in os.listdir('.'):
    if file.endswith('.pdf'):
        with open(file, 'rb') as f:
            text += pdfminer.high_level.extract_text(f)
        current_num_pages = len(list(extract_pages(file)))
        num_pages += current_num_pages
        file_names.extend([file] * current_num_pages)
        os.rename(file, os.path.join('done', file))

pages = text.split('\x0c')

def is_korekta(page_text):
    return not (re.search(fak_NUMER, page_text) and re.search(fak_DATA_WYSTAWIENIA, page_text) and 
                re.search(fak_TERMIN_PLATNOSCI, page_text) and re.search(fak_REZERWACJA, page_text))

def parse_page_data(page_text, file_name, page_number):
    numer = re.search(fak_NUMER, page_text).group(1)
    data_wystawienia = re.search(fak_DATA_WYSTAWIENIA, page_text).group(1)
    termin_platnosci = re.search(fak_TERMIN_PLATNOSCI, page_text).group(1)
    rezerwacja = int(re.search(fak_REZERWACJA, page_text).group(1))
    kwoty = re.search(fak_KWOTY, page_text, re.DOTALL).group(1).replace(' ', '').replace('\n', ' ').replace('%', '% ').strip()
    kwota_netto, stawka_vat, _, kwota_brutto = [float(kw.replace(',', '.')) if i != 1 else kw for i, kw in enumerate(kwoty.split())]

    return (numer, data_wystawienia, termin_platnosci, kwota_netto, stawka_vat, kwota_brutto, rezerwacja, file_name, page_number)

current_page_number = 0

for i, page in enumerate(pages[:len(file_names)]):
    if i >= len(file_names):
        continue
    if i > 0 and file_names[i] != file_names[i-1]:
        current_page_number = 1
    else:
        current_page_number += 1
    if is_korekta(page):
        korekty.append([file_names[i], current_page_number])
        print('korekta, ', 'plik: ', file_names[i], 'strona: ', current_page_number)
    else:
        try:
            faktura = parse_page_data(page, file_names[i], current_page_number)
            if faktura and faktura[0] not in [f[0] for f in faktury]:
                faktury.append(faktura)
            else:
                print('pominięto duplikat, ', faktura[0], 'plik: ', file_names[i], 'strona: ', current_page_number)
        except (AttributeError, ValueError) as e:
            if i != num_pages:
                print('błąd - do sprawdzenia, plik: ', file_names[i], 'strona: ', current_page_number)
            faktury.append(['błąd', 'błąd', 'błąd', 'błąd', 'błąd', 'błąd', 'błąd', file_names[i], current_page_number])


if len(faktury) > 1:
    fak = pd.DataFrame(faktury)
    kor = pd.DataFrame(korekty)
    with pd.ExcelWriter('faktury_wakacje.xlsx', mode='w') as writer:
        fak.to_excel(writer, sheet_name='faktury', index=False, header=False)
        kor.to_excel(writer, sheet_name='korekty', index=False, header=False)

columns = ['pdf_rez', 'pdf_nr_dok', 'pdf_data_dok', 'samo_start_date', 'samo_country', 'pdf_netto', 'pdf_brutto', 'pdf_termin', 'samo_date*', 'samo_adjustment*', 'pdf_st_vat', 'samo_rez', 'samo_com', 'samo_debt', 'diff']
rows = [ 
    {
        'pdf_rez': row[6],
        'pdf_nr_dok': row[0],
        'pdf_data_dok': row[1],
        'pdf_netto': row[3],
        'pdf_brutto': row[5],
        'pdf_st_vat': row[4],
        'pdf_termin': row[2],
        'samo_start_date': None,
        'samo_country': None,
        'samo_rez': None,
        'samo_com': None,
        'samo_debt': None,
        'samo_date*': None,
        'samo_adjustment*': None,
        'diff': None
    }
    for row in faktury[1:]
]

new_faktury = pd.DataFrame(rows)[columns]
with pd.ExcelWriter('faktury_wakacje.xlsx', mode='a', engine='openpyxl') as writer:
    new_faktury.to_excel(writer, sheet_name='analiza', index=False)

wb = load_workbook('faktury_wakacje.xlsx')
ws = wb['analiza']

pdf_fill = PatternFill(start_color='ADD8E6', end_color='ADD8E6', fill_type='solid')
samo_fill = PatternFill(start_color='FFA07A', end_color='FFA07A', fill_type='solid')
diff_fill = PatternFill(start_color='DDA0DD', end_color='DDA0DD', fill_type='solid')

for col in ws.iter_cols(min_row=1, max_row=1, min_col=1, max_col=len(columns)):
    for cell in col:
        if cell.value.startswith('pdf'):
            cell.fill = pdf_fill
        elif cell.value.startswith('samo'):
            cell.fill = samo_fill
        elif cell.value == 'diff':
            cell.fill = diff_fill
        
wb.save('faktury_wakacje.xlsx')
input('Press Enter to exit...')