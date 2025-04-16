#!/usr/bin/env python
# coding: utf-8

# # Analizy

# In[15]:


import logging
import pandas as pd
import re
import datetime
from collections import deque
from openpyxl.styles import PatternFill

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

# Colors for highlighting
GREEN_FILL = PatternFill(start_color="77DD76", end_color="77DD76", fill_type="solid")
GREEN_FILL2 = PatternFill(start_color="BDE7BD", end_color="BDE7BD", fill_type="solid")
RED_FILL = PatternFill(start_color="FFB6B3", end_color="FFB6B3", fill_type="solid")
RED2_FILL = PatternFill(start_color="FF6962", end_color="FF6962", fill_type="solid")


# ## Santander

# ### load_transactions

# In[16]:


def load_transactions(file_name, currency):
    df = pd.read_excel(file_name, dtype={'Kwota': float})
    df = df[df['Waluta'] == currency]
    df['kurs'] = None
    df['Kwota'] = round(df['Kwota']*100).astype(int)
    df['Data księgowania'] = pd.to_datetime(df['Data księgowania'], format='%Y-%m-%d').dt.date
    return df


# ### download_rate_from_nbp

# In[17]:


# def download_rate_from_nbp(cur, date):
#     # data is saved in tabela.csv
#     if os.path.exists('tabela.csv'):
#         logging.info('File exists')
#         df = pd.read_csv('tabela.csv', delimiter=';')
#         df['data'] = pd.to_datetime(df['data'], format='%Y%m%d')
#         df['data'] = df['data'].dt.date
#         df[f'{cur}'] = df[f'{cur}'].str.replace(',', '.').astype(float)
#         row = df[df['data'] < date].iloc[-1]
#         return row[f'{cur}']


# In[18]:


def download_rate_from_nbp(cur, date):
    startdate = date - datetime.timedelta(days=7)
    startdate = startdate.strftime('%Y-%m-%d')
    enddate = date.strftime('%Y-%m-%d')
    url = f'https://api.nbp.pl/api/exchangerates/rates/A/{cur}/{startdate}/{enddate}/'
    df = pd.read_json(url)
    df = pd.json_normalize(df['rates'])
    df['effectiveDate'] = pd.to_datetime(df['effectiveDate']).dt.date
    row = df[df['effectiveDate'] < date].iloc[-1]
    return(row['mid'])


# ### manual

# In[19]:


def manual(row, amount, balances, split_rows):
    if row['Data księgowania'] == datetime.date(2024, 10, 11) and row['Kwota'] == 43882320 and row['Tytuł'] == '00171605241011000308 - Transakcja eFX kurs: 1.0916000' and row['Waluta'] == 'USD':
        split_rows.append(pd.Series({'Kwota': 315899, 'Waluta': row['Waluta'], 'kurs': '3,9427 (z EUR)'}))
        balances.append((315899, 3.9427))
        split_rows.append(pd.Series({'Kwota': 43566421, 'Waluta': row['Waluta'], 'kurs': '3,9418 (z EUR)'}))
        balances.append((43566421, 3.9418))
    elif row['Data księgowania'] == datetime.date(2024, 10, 15) and row['Kwota'] == 153340 and row['Tytuł'] == '00171605241015000312 - Transakcja eFX kurs: 1.0905000' and row['Waluta'] == 'EUR':
        row['kurs'] = '4,2986 (z USD)'
        balances.append((153340, 4.2986))
    elif row['Data księgowania'] == datetime.date(2024, 10, 15) and row['Kwota'] == 30000000 and row['Tytuł'] == '00171605241015000314 - Transakcja eFX kurs: 1.0894000' and row['Waluta'] == 'USD':
        row['kurs'] = '3,9366 (z EUR)'
        balances.append((30000000, 3.9366))
    elif row['Data księgowania'] == datetime.date(2024, 10, 15) and row['Kwota'] == 2683334 and row['Tytuł'] == '00171605241015000315 - Transakcja eFX kurs: 1.0899000' and row['Waluta'] == 'USD':
        split_rows.append(pd.Series({'Kwota': 97427, 'Waluta': row['Waluta'], 'kurs': '3,938 (z EUR)'}))
        balances.append((97427, 3.938))
        split_rows.append(pd.Series({'Kwota': 43596, 'Waluta': row['Waluta'], 'kurs': '3,9385 (z EUR)'}))
        balances.append((43596, 3.9385))
        split_rows.append(pd.Series({'Kwota': 2542311, 'Waluta': row['Waluta'], 'kurs': '3,9387 (z EUR)'}))
        balances.append((2542311, 3.9387))
    elif row['Data księgowania'] == datetime.date(2024, 10, 18) and row['Kwota'] == 32541000 and row['Tytuł'] == '00171605241018000320 - Transakcja eFX kurs: 1.0847000' and row['Waluta'] == 'USD':
        row['kurs'] = '3,9601 (z EUR)'
        balances.append((32541000, 3.9601))
    elif row['Data księgowania'] == datetime.date(2024, 10, 31) and row['Kwota'] == 34736000 and row['Tytuł'] == '00171605241031000332 - Transakcja eFX kurs: 1.0855000' and row['Waluta'] == 'USD':
        split_rows.append(pd.Series({'Kwota': 8697841, 'Waluta': row['Waluta'], 'kurs': '4,0104 (z EUR)'}))
        balances.append((8697841, 4.0104))
        split_rows.append(pd.Series({'Kwota': 26038159, 'Waluta': row['Waluta'], 'kurs': '3,9936 (z EUR)'}))
        balances.append((26038159, 3.9936))
    elif row['Data księgowania'] == datetime.date(2024, 12, 27) and row['Kwota'] == 20806000 and row['Waluta'] == 'USD':
        row['kurs'] = '4,1083 (z EUR)'
        balances.append((20806000, 4.1083))
    elif row['Data księgowania'] == datetime.date(2024, 12, 27) and row['Kwota'] == 31100000 and row['Waluta'] == 'USD':
        split_rows.append(pd.Series({'Kwota': 27006198, 'Waluta': row['Waluta'], 'kurs': '4,1020 (z EUR)'}))
        balances.append((27006198, 4.1020))
        split_rows.append(pd.Series({'Kwota': 4093802, 'Waluta': row['Waluta'], 'kurs': '4,1057 (z EUR)'}))
        balances.append((4093802, 4.1057))
    else:
        row['kurs'] = '????'
        balances.append((amount, '????'))
    return row, balances, split_rows


# ### process_transaction

# In[20]:


def process_transaction(row, cur, balances, all_transactions):
    # Convert amount to integer (cents)
    amount = row['Kwota']
    row['kwota_pln'] = 0
    mapping = {'USD': '$', 'EUR': '€'}
    c = mapping[cur]

    original_row = row.copy()
    original_row['kurs'] = None

    # Find exchange rate in the title using regex
    match = re.search(r'eFX kurs: ([34]\.\d{4})', row['Tytuł'])
    rate = float(match.group(1)) if match else None

    if amount > 0:  # Inflow
        split_rows = []
        if rate is None:
            logging.warning(f"{row['Data księgowania']}: No exchange rate found for transaction: {row['Tytuł']}")
            if row['Rachunek strony transakcji'] not in ['76 1090 1841 0000 0001 5424 2475', '64 1090 1841 0000 0001 5424 2497', '36 1090 1841 0000 0001 5424 2516']:
                kurs = download_rate_from_nbp(cur, row['Data księgowania'])
                if balances and balances[-1][1] == kurs:
                    balances[-1] = (balances[-1][0] + amount, kurs)
                else:
                    balances.append((amount, kurs))
                row['kurs'] = str(kurs).replace('.', ',') + ' (NBP)'
            else:
                row, balances, split_rows = manual(row, amount, balances, split_rows)
        else:
            kurs = rate
            if balances and balances[-1][1] == rate:
                balances[-1] = (balances[-1][0] + amount, rate)
            else:
                balances.append((amount, rate))
            row['kurs'] = kurs
            row['kwota_pln'] = int(round(amount * rate))
        if not split_rows:
            all_transactions.append(row)
            row['saldo'] = int(round(sum(amount for amount, _ in balances)))
            if balances:
                _, rates = zip(*balances)
            else:
                rates = [0]
            row['saldo pln'] = int(round(sum(amount * rate for amount, rate in balances))) if all(r != '????' for r in rates) else '????'
            row['saldo_sklad'] = ''.join(list(f'{a/100}{c} po {b}; ' for a, b in balances))
            row['kwota_pln'] = row['Kwota'] * float(str(row['kurs']).strip(' (z EUR)').strip(' (z USD)').strip(' (NBP)').replace(',', '.')) if row['kurs'] != '????' else '????'
        else:
            all_transactions.append(original_row)
            all_transactions[-1]['saldo'] = sum(amount for amount, _ in balances)
            _, rates = zip(*balances)
            all_transactions[-1]['saldo pln'] = int(round(sum(amount * rate for amount, rate in balances))) if all(r != '????' for r in rates) else '????'
            all_transactions[-1]['saldo_sklad'] = ''.join(list(f'{a/100}{c} po {b}; ' for a, b in balances))
            all_transactions[-1]['kwota_pln'] = sum(int(round(row['Kwota'] * float(str(row['kurs']).strip(' (z EUR)').strip(' (z USD)').strip(' (NBP)').replace(',', '.')))) for row in split_rows)
            for row in split_rows:
                all_transactions.append(row)
        logging.info(f"Incoming amount: {amount}, rate: {row['kurs']}, balance state: {balances}")
    else:  # Outflow
        zejscie = -amount
        jedyne = True
        split_rows = []

        while zejscie > 0 and balances:
            mamy, kurs = balances.popleft()

            if zejscie < mamy and jedyne:
                balances.appendleft((mamy - zejscie, kurs))
                row['kurs'] = kurs
                zejscie = 0
                logging.info(f"Outgoing full amount: {zejscie}, rate: {kurs}, balance state: {balances}")
            elif zejscie < mamy and not jedyne:
                balances.appendleft((mamy - zejscie, kurs))
                split_rows.append(pd.Series({'Kwota': -zejscie, 'Waluta': row['Waluta'], 'kurs': kurs}))
                zejscie = 0
                logging.info(f"Outgoing last amount: {zejscie}, rate: {kurs}, balance state: {balances}")
            elif zejscie > mamy:
                split_row = pd.Series({'Kwota': -mamy, 'Waluta': row['Waluta'], 'kurs': kurs})
                split_rows.append(split_row)
                zejscie -= mamy
                logging.info(f"Outgoing partial amount: {mamy}, rate: {kurs}, balance state: {balances}")
                jedyne = False
            else:
                row['kurs'] = kurs
                zejscie = 0
                logging.info(f"Outgoing 1-1 amount: {mamy}, rate: {kurs}, balance state: {balances}")

        if balances:
            amounts, rates = zip(*balances)
        else:
            amounts, rates = [0], [0]

        if len(split_rows) == 1:
            print(split_rows)

        if split_rows:
            all_transactions.append(original_row)
            all_transactions[-1]['saldo'] = sum(amounts)
            all_transactions[-1]['saldo pln'] = int(round(sum(amount * rate for amount, rate in zip(amounts, rates)))) if all(r != '????' for r in rates) else '????'
            all_transactions[-1]['saldo_sklad'] = ''.join(list(f'{a/100}{c} po {b}; ' for a, b in zip(amounts, rates)))
            all_transactions[-1]['kwota_pln'] = sum(int(round(row['Kwota'] * float(str(row['kurs']).strip(' (z EUR)').strip(' (z USD)').strip(' (NBP)')))) for row in split_rows if row['kurs'] != '????')
            for row in split_rows:
                all_transactions.append(row)
        else:
            row['saldo'] = sum(amount for amount in amounts)
            row['saldo pln'] = int(round(sum(amount * rate for amount, rate in zip(amounts, rates)))) if all(r != '????' for r in rates) else '????'
            row['saldo_sklad'] = ''.join(list(f'{a/100}{c} po {b}; ' for a, b in zip(amounts, rates)))
            row['kwota_pln'] = int(round(amount * kurs)) if kurs != '????' else '????'
            all_transactions.append(row)


# ### save_transactions

# In[21]:


def save_transactions(file_name, transactions):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Create a new workbook
    wb = Workbook()
    ws = wb.active

    # convert 'Data księgowania' to Excel date format, without time
    transactions['Data księgowania'] = pd.to_datetime(transactions['Data księgowania']).dt.date
    transactions['Kwota'] = round(transactions['Kwota'] / 100, 2)
    transactions['saldo'] = round(transactions['saldo'] / 100, 2)
    transactions['saldo pln'] = pd.to_numeric(transactions['saldo pln'], errors='coerce')
    transactions['saldo pln'] = round(transactions['saldo pln'] / 100, 2)
    transactions['kwota_pln'] = pd.to_numeric(transactions['kwota_pln'], errors='coerce')
    transactions['kwota_pln'] = round(-transactions['kwota_pln'] / 100, 2)
    
    # Write data to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(transactions, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Highlight rows based on amount
            if r_idx > 1:  # Skip header
                if transactions.iloc[r_idx - 2]['Kwota'] > 0:
                    if pd.isna(transactions.iloc[r_idx - 2]['Data księgowania']):
                        cell.fill = GREEN_FILL2
                    else:
                        cell.fill = GREEN_FILL
                elif transactions.iloc[r_idx - 2]['Kwota'] < 0:
                    if pd.isna(transactions.iloc[r_idx - 2]['Data księgowania']):
                        cell.fill = RED_FILL
                    else:
                        cell.fill = RED2_FILL

    wb.save(file_name)


# ### move_row_up

# In[22]:


def move_row_up(transactions, cur, target_date, target_amount, positions=2):
    target_date = pd.Timestamp(target_date)  # Ensure consistent date format

    logging.info(f"Target date: {target_date}, Target amount: {target_amount}, Target currency: {cur}")

    # Log unique values in columns involved in the condition
    logging.debug(f"Unique 'Data księgowania':\n{transactions['Data księgowania'].unique()}")
    logging.debug(f"Unique 'Kwota':\n{transactions['Kwota'].unique()}")
    logging.debug(f"Unique 'Waluta':\n{transactions['Waluta'].unique()}")

    # Define the condition
    condition = (
        (transactions['Data księgowania'] == target_date) &
        (transactions['Kwota'] == target_amount) &
        (transactions['Waluta'] == cur)
    )

    # Find matching rows
    idx = transactions[condition].index

    if idx.empty:
        logging.warning("No rows match the condition.")
        logging.debug(f"Condition details:\n"
                      f"Target date matches:\n{transactions['Data księgowania'] == target_date}\n"
                      f"Target amount matches:\n{transactions['Kwota'] == target_amount}\n"
                      f"Target currency matches:\n{transactions['Waluta'] == cur}")
    else:
        logging.info(f"Matched row index: {idx}")

    # If a match exists, move the row
    if not idx.empty:
        idx = idx[0]
        if idx >= positions:
            logging.info(f"Moving row at index {idx} up by {positions} positions.")
            row_to_move = transactions.iloc[idx]
            transactions = pd.concat([
                transactions.iloc[:idx - positions],
                pd.DataFrame([row_to_move]),
                transactions.iloc[idx - positions:idx],
                transactions.iloc[idx + 1:]
            ]).reset_index(drop=True)
        else:
            logging.warning(f"Cannot move row at index {idx} up by {positions}. Not enough rows above.")
    return transactions


# ### main

# In[23]:


def main():
    for cur in ['USD', 'EUR']:
        balances = deque()
        all_transactions = []
        input_file = 'san.xlsx'
        transactions = load_transactions(input_file, cur)
        transactions.sort_values(by=['Data księgowania', 'Kwota'], ascending=[True, False], inplace=True, ignore_index=True)
        # transactions = transactions[transactions['Data księgowania'] <= datetime.date(2024, 12, 16)]

        # move the target row
        transactions = move_row_up(transactions, cur, datetime.date(2024, 10, 15), -167217, 2)

        # Process each transaction
        for _, row in transactions.iterrows():
            process_transaction(row, cur, balances, all_transactions)

        date = transactions['Data księgowania'].max()
        output_file = f"Santander {cur} do {date}.xlsx"


        # Save transactions to another file
        save_transactions(output_file, pd.DataFrame(all_transactions))

if __name__ == "__main__":
    main()


# ## BNP

# ### load_transactions_bnp

# In[24]:


def load_transactions_bnp(file_name, currency):
    df = pd.read_csv(file_name, delimiter=';', header=None, names=['data2', 'data', 'typ', 'kto', 'tytuł', 'kwota', 'waluta', 'kod'], index_col=False, encoding='windows-1250')
    df = df[df['waluta'] == currency]
    df['kurs'] = None
    df['kwota'] = df['kwota'].str.replace(',', '.').astype(float)
    df['kwota'] = round(df['kwota']*100).astype(int)
    df['data'] = pd.to_datetime(df['data'], format='%d.%m.%Y').dt.date
    return df


# In[25]:


def manual_bnp(row, balances, split_rows):
    if row['data'] == datetime.date(2024, 11, 13) and row['kwota'] == 727696:
        logging.info('NADPISYWANIE')
        split_rows.append(pd.Series({'kwota': 529382, 'waluta': 'USD', 'kurs': '4,0435 (z EUR)'}))
        balances.append((529382, 4.0435))
        split_rows.append(pd.Series({'kwota': 47661, 'waluta': 'USD', 'kurs': '4,0443 (z EUR)'}))
        balances.append((47661, 4.0443))
        split_rows.append(pd.Series({'kwota': 32904, 'waluta': 'USD', 'kurs': '4,0431 (z EUR)'}))
        balances.append((32904, 4.0431))
        split_rows.append(pd.Series({'kwota': 43517, 'waluta': 'USD', 'kurs': '4,0842 (z EUR)'}))
        balances.append((43517, 4.0842))
        split_rows.append(pd.Series({'kwota': 74232, 'waluta': 'USD', 'kurs': '4,0742 (z EUR)'}))
        balances.append((74232, 4.0742))
    elif row['data'] == datetime.date(2024, 12, 27) and row['kwota'] == 416520:
        split_rows.append(pd.Series({'kwota': 303074, 'waluta': 'USD', 'kurs': '4,1038 (z EUR)'}))
        balances.append((303073, 4.1038))
        split_rows.append(pd.Series({'kwota': 113446, 'waluta': 'USD', 'kurs': '4,0884 (z EUR)'}))
        balances.append((113447, 4.0884))
    elif row['data'] == datetime.date(2025, 2, 20) and row['kwota'] == 200000:
        row['kurs'] = '4,1798 (z USD)'
        balances.append((200000, 4.1798))
    elif row['data'] == datetime.date(2025, 4, 9) and row['kwota'] == 13500000:
        row['kurs'] = '4,2124 (z USD)'
        balances.append((13500000, 4.2124))
    else:
        row['kurs'] = '????'

    return row, balances, split_rows


# In[26]:


def process_transaction_bnp(row, cur, balances, all_transactions):
    amount = row['kwota']
    row['kwota_pln'] = 0
    mapping = {'USD': '$', 'EUR': '€'}
    c = mapping[cur]

    original_row = row.copy()
    original_row['kurs'] = None

    title = str(row['tytuł']) if pd.notna(row['tytuł']) else ""
    if 'BO' in title:
        print (title)
        print(re.search(r'(EUR|USD) PLN ([34]\.\d{4})', title))


    match = re.search(r'(EUR|USD) PLN ([34]\.\d{4})', title)
    rate = float(match.group(2)) if match else None

    kurs = None

    if amount > 0:  # Inflow
        if row['data'] == datetime.date(2024, 12, 23) and row['kwota'] == 27000000:
            print('NADPISYWANIE')
            row['kurs'] = 4.1071
            row['saldo'] = 26988840
            row['saldo pln'] = 26988840 * 4.1071
            row['saldo_sklad'] = '269888,40$ po 4,1071; '
            row['kwota_pln'] = 27000000 * 4.1071
            balances.appendleft((26988840, 4.1071))
            all_transactions.append(row)
        else:
            split_rows = []
            if rate is None:
                logging.warning(f"{row['data']} - No exchange rate found for transaction: {row['tytuł'], row['kwota']}")
                if 'USD EUR' not in title and 'EUR USD' not in title:
                    kurs = download_rate_from_nbp(cur, row['data'])
                    if balances and balances[-1][1] == kurs:
                        balances[-1] = (balances[-1][0] + amount, kurs)
                    else:
                        balances.append((amount, kurs))
                    row['kurs'] = str(kurs).replace('.', ',') + ' (NBP)'
                else:
                    original_row, balances, split_rows = manual_bnp(row, balances, split_rows)
            else:
                kurs = rate
                row['kurs'] = rate
                if balances and balances[-1][1] == rate:
                    balances[-1] = (balances[-1][0] + amount, rate)
                else:
                    balances.append((amount, rate))
                try:
                    if isinstance(row['kurs'], str):
                        kurs = float(row['kurs'].replace(',', '.').replace(' (NBP)', '').strip())
                    elif isinstance(row['kurs'], float):
                        kurs = row['kurs']
                except ValueError:
                    logging.error(f"Error converting kurs: {row['kurs']}")
                    kurs = None

            if not split_rows:
                all_transactions.append(row)
                row['saldo'] = int(round(sum(amount for amount, _ in balances)))
                if balances:
                    _, rates = zip(*balances)
                else:
                    rates = [0]
                row['saldo pln'] = int(round(sum(amount * rate for amount, rate in balances))) if all(r != '????' for r in rates) else '????'
                row['saldo_sklad'] = ''.join(list(f'{a/100}{c} po {b}; ' for a, b in balances))

                print(f"Debug: kurs = {row['kurs']}, amount = {amount}, type(kurs) = {type(row['kurs'])}")

                print(row['kurs'])
                print(kurs, type(kurs), amount, type(amount)) if kurs else print('????')
                try:
                    cleaned_kurs = float(str(row['kurs']).strip(' (z EUR)').strip(' (z USD)').strip(' (NBP)').replace(',', '.'))
                    row['kwota_pln'] = int(round(amount * cleaned_kurs))
                    print(f"Debug: cleaned_kurs = {cleaned_kurs}, kwota_pln = {row['kwota_pln']}")
                except ValueError as e:
                    print(f"Error converting kurs: {e}")
                    row['kwota_pln'] = '????'
                # row['kwota_pln'] = int(round(amount * float(str(row['kurs']).strip(' (z EUR)').strip(' (z USD)').strip(' (NBP)').replace(',', '.')))) if row['kurs'] != '????' and row['kurs'] else '????'
            else:
                all_transactions.append(original_row)
                all_transactions[-1]['saldo'] = sum(amount for amount, _ in balances)
                _, rates = zip(*balances)
                all_transactions[-1]['saldo pln'] = int(round(sum(amount * rate for amount, rate in balances))) if all(r != '????' for r in rates) else '????'
                all_transactions[-1]['saldo_sklad'] = ''.join(list(f'{a/100}{c} po {b}; ' for a, b in balances))
                all_transactions[-1]['kwota_pln'] = sum(int(round(row['kwota'] * float(str(row['kurs']).strip(' (z EUR)').strip(' (z USD)').strip(' (NBP)').replace(',', '.')))) for row in split_rows)
                print(all_transactions[-1]['kwota_pln'])
                for row in split_rows:
                    all_transactions.append(row)
            logging.info(f"Incoming amount: {amount}, rate: {row['kurs']}, balance state: {balances}")
    else:  # Outflow
        split_rows = []
        if row['data'] == datetime.date(2024, 12, 20) and row['kwota'] == -13066000:
            print('NADPISYWANIE')
            split_rows.append(pd.Series({'kwota': -69930, 'waluta': 'USD', 'kurs': '4,0915 (z EUR)'}))
            split_rows.append(pd.Series({'kwota': -12984910, 'waluta': 'USD', 'kurs': '4,0944 (z EUR)'}))
            split_rows.append(pd.Series({'kwota': -11160, 'waluta': 'USD', 'kurs': '4,1071'}))
            balances.pop()
            balances.pop()
            all_transactions.append(row)
            all_transactions[-1]['saldo'] = -11160
            all_transactions[-1]['saldo pln'] = -11160 * 4.1071
            all_transactions[-1]['saldo_sklad'] = '-111,60$ po 4,1071; '
            all_transactions[-1]['kwota_pln'] = -69930 * 4.0915 + -12984910 * 4.0944 + -11160 * 4.1071
            all_transactions.append(split_rows[0])
            all_transactions.append(split_rows[1])
            all_transactions.append(split_rows[2])
        else:
            zejscie = -amount
            jedyne = True

            while zejscie > 0 and balances:
                mamy, kurs = balances.popleft()

                if zejscie < mamy and jedyne:
                    balances.appendleft((mamy - zejscie, kurs))
                    row['kurs'] = kurs
                    zejscie = 0
                    logging.info(f"Outgoing full amount: {zejscie}, rate: {kurs}, balance state: {balances}")
                elif zejscie < mamy and not jedyne:
                    balances.appendleft((mamy - zejscie, kurs))
                    split_rows.append(pd.Series({'kwota': -zejscie, 'waluta': row['waluta'], 'kurs': kurs}))
                    zejscie = 0
                    logging.info(f"Outgoing last amount: {zejscie}, rate: {kurs}, balance state: {balances}")
                elif zejscie > mamy:
                    split_row = pd.Series({'kwota': -mamy, 'waluta': row['waluta'], 'kurs': kurs})
                    split_rows.append(split_row)
                    zejscie -= mamy
                    logging.info(f"Outgoing partial amount: {mamy}, rate: {kurs}, balance state: {balances}")
                    jedyne = False
                else:
                    row['kurs'] = kurs
                    zejscie = 0
                    logging.info(f"Outgoing 1-1 amount: {mamy}, rate: {kurs}, balance state: {balances}")

            if balances:
                amounts, rates = zip(*balances)
            else:
                amounts, rates = [0], [0]

            if len(split_rows) == 1:
                print(split_rows)
            
            if split_rows:
                all_transactions.append(original_row)
                all_transactions[-1]['saldo'] = sum(amounts)
                all_transactions[-1]['saldo pln'] = int(round(sum(amount * rate for amount, rate in zip(amounts, rates)))) if all(r != '????' for r in rates) else '????'
                all_transactions[-1]['saldo_sklad'] = ''.join(list(f'{a/100}{c} po {b}; ' for a, b in zip(amounts, rates)))
                all_transactions[-1]['kwota_pln'] = sum(int(round(row['kwota'] * float(str(row['kurs']).strip(' (z EUR)').strip(' (z USD)').strip(' (NBP)').replace(',', '.')))) for row in split_rows if row['kurs'] != '????')
                for row in split_rows:
                    all_transactions.append(row)
            else:
                row['saldo'] = sum(amount for amount in amounts)
                row['saldo pln'] = int(round(sum(amount * rate for amount, rate in zip(amounts, rates)))) if all(r != '????' for r in rates) else '????'
                row['saldo_sklad'] = ''.join(list(f'{a/100}{c} po {b}; ' for a, b in zip(amounts, rates)))
                row['kwota_pln'] = int(round(amount * kurs)) if kurs != '????' else '????'
                all_transactions.append(row)

    return all_transactions


# In[27]:


def save_transactions_bnp(file_name, transactions):
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows

    # Create a new workbook
    wb = Workbook()
    ws = wb.active

    # convert 'Data księgowania' to Excel date format, without time
    transactions['data'] = pd.to_datetime(transactions['data']).dt.date
    transactions['kwota'] = round(transactions['kwota'] / 100, 2)
    transactions['saldo'] = round(transactions['saldo'] / 100, 2)
    transactions['saldo pln'] = round(transactions['saldo pln'] / 100, 2)
    transactions['kwota_pln'] = pd.to_numeric(transactions['kwota_pln'], errors='coerce')
    transactions['kwota_pln'] = round(-transactions['kwota_pln'] / 100, 2)
    
    # Write data to the worksheet
    for r_idx, row in enumerate(dataframe_to_rows(transactions, index=False, header=True), start=1):
        for c_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            # Highlight rows based on amount
            if r_idx > 1:  # Skip header
                if transactions.iloc[r_idx - 2]['kwota'] > 0:
                    if pd.isna(transactions.iloc[r_idx - 2]['typ']):
                        cell.fill = GREEN_FILL2
                    else:
                        cell.fill = GREEN_FILL
                elif transactions.iloc[r_idx - 2]['kwota'] < 0:
                    if pd.isna(transactions.iloc[r_idx - 2]['typ']):
                        cell.fill = RED_FILL
                    else:
                        cell.fill = RED2_FILL

    wb.save(file_name)


# In[28]:


def main():
    for cur in ['USD', 'EUR']:
        balances = deque()
        all_transactions = []
        input_file = 'bnp.csv'
        transactions = load_transactions_bnp(input_file, cur)
        transactions.sort_values(by=['data', 'kwota'], ascending=[True, False], inplace=True, ignore_index=True)

        # process each transaction
        for _, row in transactions.iterrows():
            all_transactions = process_transaction_bnp(row, cur, balances, all_transactions)

        date = transactions['data'].max()
        output_file = f"BNP {cur} do {date}.xlsx"
        
        # save transactions to another file
        save_transactions_bnp(output_file, pd.DataFrame(all_transactions))

if __name__ == "__main__":
    main()


# In[ ]:




